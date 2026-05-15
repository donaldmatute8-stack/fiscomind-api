#!/usr/bin/env python3
"""
FiscoMind Railway API v4.1 - Based on working local version
- Same SAT calls as local api_server.py
- METADATA + EstadoComprobante.TODOS
- recover_comprobante_issued_request (not emitted)
- Parse base64→ZIP→CSV (METADATA format)
"""

import os
import sys
import json
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict

from flask import Flask, jsonify, request
from flask_cors import CORS

sys.path.insert(0, str(Path(__file__).parent / "src"))

app = Flask(__name__)
CORS(app, origins=["*"], supports_credentials=True)

import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("fiscomind-api")

DATA_DIR = Path(os.environ.get("DATA_DIR", "/app/data"))
DATA_DIR.mkdir(exist_ok=True, parents=True)
CACHE_FILE = DATA_DIR / "cfdi_cache.json"

MONTH_NAMES = {
    "01": "Enero",
    "02": "Febrero",
    "03": "Marzo",
    "04": "Abril",
    "05": "Mayo",
    "06": "Junio",
    "07": "Julio",
    "08": "Agosto",
    "09": "Septiembre",
    "10": "Octubre",
    "11": "Noviembre",
    "12": "Diciembre",
}

# ─── SAT Connector (lazy) ──────────────────────────────────────────

_sat_connector = None


def sat_connector():
    global _sat_connector
    if _sat_connector is None:
        try:
            from sat_connector_real import SATConnectorReal

            _sat_connector = SATConnectorReal(rfc="MUTM8610091NA")
            logger.info(
                f"SAT connector created. Vault: {_sat_connector.vault.VAULT_DIR}"
            )
        except Exception as e:
            logger.error(f"SAT connector error: {e}", exc_info=True)
    return _sat_connector


# ─── Cache ──────────────────────────────────────────────────────────


def load_cache():
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text())
        except:
            pass
    return {"last_sync": None, "recibidos": [], "emitidos": [], "pending_requests": {}}


def save_cache(data):
    CACHE_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str))


# ─── Routes ─────────────────────────────────────────────────────────


@app.route("/")
def index():
    return jsonify(
        {
            "name": "FiscoMind API",
            "version": "4.1.0",
            "status": "running",
            "mode": "REAL",
            "endpoints": {
                "GET /": "Info",
                "GET /health": "Health check",
                "GET /dashboard": "Dashboard fiscal (datos reales)",
                "POST /sync": "Submit CFDI download request (async)",
                "POST /sync/check": "Check pending requests & download",
                "GET /cfdis": "Lista CFDIs",
                "GET /emitidos": "Facturas emitidas",
                "POST /emitir": "Emitir factura CFDI 4.0",
                "POST /cancelar": "Cancelar factura",
                "POST /sync": "Submit CFDI download request (async)",
                "POST /sync/check": "Check pending requests & download",
                "GET /obligaciones": "Obligaciones fiscales",
                "GET /opinion": "Opinión de cumplimiento SAT",
                "GET /optimize/suggestions": "Sugerencias de optimización fiscal",
                "GET /optimize/projection": "Proyección ISR + detección salto de tarifa",
            },
        }
    )


@app.route("/health")
def health():
    return jsonify(
        {"status": "ok", "timestamp": datetime.now().isoformat(), "mode": "REAL"}
    )


@app.route("/dashboard")
def dashboard():
    cache = load_cache()
    recibidos = cache.get("recibidos", [])
    emitidos = cache.get("emitidos", [])

    total_ingresos = sum(
        float(c.get("monto", 0))
        for c in recibidos
        if c.get("efecto") == "I" and c.get("estatus") == "1"
    )
    total_egresos = sum(
        float(c.get("monto", 0))
        for c in recibidos
        if c.get("efecto") == "E" and c.get("estatus") == "1"
    )
    total_emitidos = sum(
        float(c.get("monto", 0)) for c in emitidos if c.get("estatus") == "1"
    )
    deducibles = [c for c in recibidos if c.get("deductible")]
    total_deducible = sum(float(c.get("monto", 0)) for c in deducibles)

    return jsonify(
        {
            "rfc": "MUTM8610091NA",
            "last_sync": cache.get("last_sync"),
            "summary": {
                "total_recibidos": len(recibidos),
                "total_emitidos": len(emitidos),
                "total_ingresos": round(total_ingresos, 2),
                "total_egresos": round(total_egresos, 2),
                "total_emitido_monto": round(total_emitidos, 2),
                "total_deducible": round(total_deducible, 2),
                "ahorro_isr_estimado": round(total_deducible * 0.30, 2),
                "deducibles_count": len(deducibles),
            },
            "recent_cfdis": sorted(
                recibidos + emitidos,
                key=lambda x: x.get("fecha_emision", ""),
                reverse=True,
            )[:20],
        }
    )


@app.route("/sync", methods=["POST"])
def sync():
    """Submit CFDI download request to SAT (returns immediately)"""
    connector = sat_connector()
    if not connector:
        return jsonify(
            {"status": "error", "message": "SAT connector no disponible"}
        ), 500

    data = request.json or {}
    date_start = data.get("date_start", (date.today() - timedelta(days=90)).isoformat())
    date_end = data.get("date_end", date.today().isoformat())
    tipo = data.get("tipo", "recibidos")

    try:
        if not connector._is_connected:
            if not connector.authenticate():
                return jsonify(
                    {"status": "error", "message": "Autenticación FIEL fallida"}
                ), 401

        result = connector.submit_download_request(date_start, date_end, tipo)

        if result.get("status") == "submitted":
            cache = load_cache()
            cache.setdefault("pending_requests", {})[result["id_solicitud"]] = {
                "id_solicitud": result["id_solicitud"],
                "tipo": tipo,
                "date_start": date_start,
                "date_end": date_end,
                "submitted": datetime.now().isoformat(),
            }
            save_cache(cache)

        return jsonify(result)

    except Exception as e:
        logger.error(f"Sync failed: {e}", exc_info=True)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/sync/check", methods=["POST"])
def sync_check():
    """Check pending requests and download CFDIs if ready"""
    connector = sat_connector()
    if not connector:
        return jsonify(
            {"status": "error", "message": "SAT connector no disponible"}
        ), 500

    data = request.json or {}
    id_solicitud = data.get("id_solicitud")

    try:
        if not connector._is_connected:
            if not connector.authenticate():
                return jsonify({"status": "error", "message": "Not authenticated"}), 401

        cache = load_cache()
        pending = cache.get("pending_requests", {})

        if id_solicitud:
            # Check specific request
            status = connector.check_request_status(id_solicitud)
            if status.get("finished"):
                cfdis = status.get("cfdis", [])
                if cfdis:
                    req_info = pending.get(id_solicitud, {})
                    key = (
                        "recibidos"
                        if req_info.get("tipo") == "recibidos"
                        else "emitidos"
                    )
                    cache.setdefault(key, []).extend(cfdis)
                    cache["last_sync"] = datetime.now().isoformat()
                    if id_solicitud in pending:
                        del pending[id_solicitud]
                    cache["pending_requests"] = pending
                    save_cache(cache)
            return jsonify(status)

        # Check all pending
        if not pending:
            return jsonify(
                {"status": "success", "message": "No pending requests", "pending": 0}
            )

        results = []
        for req_id in list(pending.keys()):
            status = connector.check_request_status(req_id)
            results.append(status)

            if status.get("finished"):
                cfdis = status.get("cfdis", [])
                if cfdis:
                    req_info = pending[req_id]
                    key = (
                        "recibidos"
                        if req_info.get("tipo") == "recibidos"
                        else "emitidos"
                    )
                    cache.setdefault(key, []).extend(cfdis)
                    cache["last_sync"] = datetime.now().isoformat()
                if req_id in pending:
                    del pending[req_id]

        cache["pending_requests"] = pending
        save_cache(cache)

        return jsonify(
            {"status": "success", "checked": len(results), "results": results}
        )

    except Exception as e:
        logger.error(f"Sync check failed: {e}", exc_info=True)
        return jsonify({"status": "error", "message": str(e)}), 500


# ─── SISTEMA DE FECHAS Y FILTROS ─────────────────────────────────────

from datetime import datetime as dt


def parse_fecha(fecha_str):
    """Parse fecha desde string ISO o DD/MM/YYYY"""
    try:
        return date.fromisoformat(fecha_str)
    except:
        try:
            return dt.strptime(fecha_str, "%d/%m/%Y").date()
        except:
            return None


def get_mes_fiscal(fecha_str):
    """Extraer mes fiscal de fecha ISO"""
    try:
        d = date.fromisoformat(fecha_str)
        return f"{d.year}-{str(d.month).zfill(2)}"
    except:
        return None


def get_trimestre(fecha_str):
    """Extraer trimestre de fecha ISO"""
    try:
        d = date.fromisoformat(fecha_str)
        q = (d.month - 1) // 3 + 1
        return f"{d.year}-Q{q}"
    except:
        return None


def calcular_isr_estimado(
    ingresos, deducciones=0, regimen="612", actividad="profesional"
):
    """
    Calculo estimado de ISR para personas físicas
    Regimen 612: Actividades Profesionales
    """
    base_gravable = max(0, ingresos - deducciones)

    # Tabla ISR 2026 (aproximada)
    if base_gravable <= 7464.05:
        isr = base_gravable * 0.0192
    elif base_gravable <= 63324.11:
        isr = base_gravable * 0.064 + 133.28
    elif base_gravable <= 111396.25:
        isr = base_gravable * 0.1088 + 3357.30
    elif base_gravable <= 129921.73:
        isr = base_gravable * 0.128 + 7976.64
    elif base_gravable <= 155229.00:
        isr = base_gravable * 0.1664 + 10364.16
    elif base_gravable <= 296737.70:
        isr = base_gravable * 0.1976 + 15356.85
    elif base_gravable <= 392842.86:
        isr = base_gravable * 0.2195 + 43631.86
    elif base_gravable <= 939929.66:
        isr = base_gravable * 0.2388 + 64769.14
    else:
        isr = base_gravable * 0.30 + 176570.65

    return {
        "ingresos": round(ingresos, 2),
        "deducciones": round(deducciones, 2),
        "base_gravable": round(base_gravable, 2),
        "isr_estimado": round(isr, 2),
        "reserva_mensual": round(isr / 12, 2),
        "tasa_efectiva": round((isr / ingresos) * 100, 2) if ingresos > 0 else 0,
    }


# Categorías de gastos deducibles (SAT)
CATEGORIAS_DEDUCIBLES = {
    "G01": {"nombre": "Adquisición de mercancías", "deducible": True, "limite": None},
    "G02": {
        "nombre": "Devoluciones, descuentos, bonificaciones",
        "deducible": True,
        "limite": None,
    },
    "G03": {"nombre": "Gastos en general", "deducible": True, "limite": None},
    "I01": {
        "nombre": "Construcciones",
        "deducible": True,
        "limite": None,
        "depreciable": True,
    },
    "I02": {
        "nombre": "Mobiliario y equipo de oficina",
        "deducible": True,
        "limite": None,
        "depreciable": True,
    },
    "I03": {
        "nombre": "Equipo de transporte",
        "deducible": True,
        "limite": None,
        "depreciable": True,
    },
    "I04": {
        "nombre": "Equipo de computo",
        "deducible": True,
        "limite": None,
        "depreciable": True,
    },
    "I05": {
        "nombre": "Dados, troqueles, moldes, matrices",
        "deducible": True,
        "limite": None,
    },
    "I06": {"nombre": "Comunicaciones telefónicas", "deducible": True, "limite": None},
    "I07": {"nombre": "Comunicaciones satelitales", "deducible": True, "limite": None},
    "I08": {"nombre": "Otra maquinaria y equipo", "deducible": True, "limite": None},
    "D01": {"nombre": "Honorarios médicos", "deducible": True, "limite": "personales"},
    "D02": {
        "nombre": "Gastos médicos por incapacidad",
        "deducible": True,
        "limite": "personales",
    },
    "D03": {"nombre": "Gastos funerales", "deducible": True, "limite": "personales"},
    "D04": {"nombre": "Donativos", "deducible": True, "limite": "personales"},
    "D05": {
        "nombre": "Intereses reales hipotecarios",
        "deducible": True,
        "limite": "personales",
    },
    "D06": {
        "nombre": "Aportaciones voluntarias SAR",
        "deducible": True,
        "limite": "personales",
    },
    "D07": {
        "nombre": "Primas de seguros de gastos médicos",
        "deducible": True,
        "limite": "personales",
    },
    "D08": {
        "nombre": "Gastos de transportación escolar",
        "deducible": True,
        "limite": "personales",
    },
    "D09": {
        "nombre": "Depósitos en cuentas de ahorro",
        "deducible": True,
        "limite": "personales",
    },
    "D10": {
        "nombre": "Servicios educativos (colegiatura)",
        "deducible": True,
        "limite": "personales",
    },
    "S01": {"nombre": "Sin efectos fiscales", "deducible": False},
    "CP01": {"nombre": "Pagos", "deducible": False},
    "CN01": {"nombre": "Nómina", "deducible": True},
}


def clasificar_gasto(cfdi):
    """Clasifica un CFDI y determina si es deducible"""
    uso_cfdi = cfdi.get("uso_cfdi", "S01")
    efecto = cfdi.get("efecto", "I")

    categoria = CATEGORIAS_DEDUCIBLES.get(
        uso_cfdi, {"nombre": "No clasificado", "deducible": False}
    )

    # Solo ingresos (efecto I) pueden ser deducibles en gastos
    # Egresos (efecto E) son tuyos emitidos

    es_deducible = categoria.get("deducible", False)

    # Si es factura recibida (efecto I) y tiene categoría deducible
    if efecto == "I" and es_deducible:
        return {
            "clasificacion": categoria["nombre"],
            "uso_cfdi": uso_cfdi,
            "deducible": True,
            "monto_deducible": float(cfdi.get("monto", 0)),
            "notas": categoria.get("limite", None),
        }
    else:
        return {
            "clasificacion": categoria["nombre"] if es_deducible else "No deducible",
            "uso_cfdi": uso_cfdi,
            "deducible": False,
            "monto_deducible": 0,
        }


@app.route("/cfdis")
def list_cfdis():
    """List CFDIs con filtros de fechas y paginación"""
    cache = load_cache()
    tipo = request.args.get("tipo", "all")

    # Filtros de fecha
    desde = request.args.get("desde")  # YYYY-MM-DD
    hasta = request.args.get("hasta")  # YYYY-MM-DD
    mes = request.args.get("mes")  # YYYY-MM
    trimestre = request.args.get("trimestre")  # YYYY-Q#
    anio = request.args.get("anio")  # YYYY

    # Filtros adicionales
    efecto = request.args.get("efecto")  # I (Ingreso), E (Egreso), P (Pago)
    estatus = request.args.get("estatus")  # 1 (Vigente), 0 (Cancelada)
    min_monto = request.args.get("min_monto", type=float)
    max_monto = request.args.get("max_monto", type=float)

    cfdis = []
    if tipo in ["all", "recibidos"]:
        cfdis.extend(cache.get("recibidos", []))
    if tipo in ["all", "emitidos"]:
        cfdis.extend(cache.get("emitidos", []))

    # Aplicar filtros
    resultado = []
    for c in cfdis:
        incluir = True
        fecha_emision = c.get("fecha_emision", "")

        # Filtro rango de fechas
        if desde:
            try:
                if date.fromisoformat(fecha_emision) < date.fromisoformat(desde):
                    incluir = False
            except:
                pass

        if hasta:
            try:
                if date.fromisoformat(fecha_emision) > date.fromisoformat(hasta):
                    incluir = False
            except:
                pass

        # Filtro por mes fiscal
        if mes and mes != get_mes_fiscal(fecha_emision):
            incluir = False

        # Filtro por trimestre
        if trimestre and trimestre != get_trimestre(fecha_emision):
            incluir = False

        # Filtro por año
        if anio and not fecha_emision.startswith(anio):
            incluir = False

        # Filtro por efecto
        if efecto and c.get("efecto") != efecto:
            incluir = False

        # Filtro por estatus
        if estatus and str(c.get("estatus")) != estatus:
            incluir = False

        # Filtro por monto
        monto = float(c.get("monto", 0))
        if min_monto is not None and monto < min_monto:
            incluir = False
        if max_monto is not None and monto > max_monto:
            incluir = False

        if incluir:
            resultado.append(c)

    # Calcular totales
    total_monto = sum(float(c.get("monto", 0)) for c in resultado)

    return jsonify(
        {
            "rfc": "MUTM8610091NA",
            "tipo": tipo,
            "total": len(resultado),
            "total_monto": round(total_monto, 2),
            "filtros_aplicados": {
                "desde": desde,
                "hasta": hasta,
                "mes": mes,
                "trimestre": trimestre,
                "anio": anio,
                "efecto": efecto,
            },
            "cfdis": resultado,
        }
    )


@app.route("/emitidos")
def list_emitidos():
    cache = load_cache()
    emitidos = cache.get("emitidos", [])
    total = sum(float(c.get("monto", 0)) for c in emitidos)
    return jsonify(
        {
            "rfc": "MUTM8610091NA",
            "total_emitidos": len(emitidos),
            "total_monto": round(total, 2),
            "cfdis": emitidos,
        }
    )


@app.route("/resumen-mensual")
def resumen_mensual():
    """Resumen fiscal por mes"""
    mes = request.args.get("mes", date.today().strftime("%Y-%m"))  # YYYY-MM
    cache = load_cache()

    try:
        year, month = mes.split("-")
        year, month = int(year), int(month)
    except:
        return jsonify({"error": "Formato de mes debe ser YYYY-MM"}), 400

    # Filtrar CFDIs del mes
    cfdis_recibidos = [
        c
        for c in cache.get("recibidos", [])
        if get_mes_fiscal(c.get("fecha_emision", "")) == mes
    ]
    cfdis_emitidos = [
        c
        for c in cache.get("emitidos", [])
        if get_mes_fiscal(c.get("fecha_emision", "")) == mes
    ]

    # Calcular totales
    ingresos = sum(
        float(c.get("monto", 0)) for c in cfdis_recibidos if c.get("efecto") == "I"
    )
    egresos = sum(
        float(c.get("monto", 0)) for c in cfdis_recibidos if c.get("efecto") == "E"
    )
    pagos = sum(
        float(c.get("monto", 0)) for c in cfdis_recibidos if c.get("efecto") == "P"
    )

    # Calcular deducciones
    deducciones = 0
    deducciones_detalle = []
    for c in cfdis_recibidos:
        clasif = clasificar_gasto(c)
        if clasif["deducible"]:
            monto = clasif["monto_deducible"]
            deducciones += monto
            deducciones_detalle.append(
                {
                    "uuid": c.get("uuid", ""),
                    "monto": monto,
                    "clasificacion": clasif["clasificacion"],
                    "uso_cfdi": clasif["uso_cfdi"],
                }
            )

    # Calcular ISR estimado
    isr = calcular_isr_estimado(ingresos, deducciones)

    return jsonify(
        {
            "mes": mes,
            "nombre_mes": f"{MONTH_NAMES.get(str(month).zfill(2), '')} {year}",
            "cfdis_recibidos": len(cfdis_recibidos),
            "cfdis_emitidos": len(cfdis_emitidos),
            "ingresos": round(ingresos, 2),
            "egresos": round(egresos, 2),
            "pagos": round(pagos, 2),
            "deducciones": round(deducciones, 2),
            "deducciones_detalle": deducciones_detalle[:10],  # Top 10
            "isr": isr,
            "estado_cuenta": {
                "saldo": round(ingresos - egresos, 2),
                "base_gravable": isr["base_gravable"],
                "reserva_isr": isr["reserva_mensual"],
            },
        }
    )


@app.route("/resumen-trimestral")
def resumen_trimestral():
    """Resumen fiscal por trimestre"""
    trimestre = request.args.get("trimestre")  # YYYY-Q#
    if not trimestre:
        # Calcular trimestre actual
        today = date.today()
        q = (today.month - 1) // 3 + 1
        trimestre = f"{today.year}-Q{q}"

    cache = load_cache()

    try:
        year, q = trimestre.split("-Q")
        year = int(year)
        q = int(q)
    except:
        return jsonify({"error": "Formato de trimestre debe ser YYYY-Q#"}), 400

    # Calcular meses del trimestre
    meses = [f"{year}-{str(m).zfill(2)}" for m in range((q - 1) * 3 + 1, q * 3 + 1)]

    # Agregar resumen de cada mes
    resumen_meses = []
    total_ingresos = 0
    total_deducciones = 0

    for mes in meses:
        # Reutilizar lógica de resumen mensual
        cfdis_mes = [
            c
            for c in cache.get("recibidos", [])
            if get_mes_fiscal(c.get("fecha_emision", "")) == mes
        ]

        ingresos_mes = sum(
            float(c.get("monto", 0)) for c in cfdis_mes if c.get("efecto") == "I"
        )
        deducciones_mes = sum(
            float(c.get("monto", 0))
            for c in cfdis_mes
            if clasificar_gasto(c)["deducible"]
        )

        total_ingresos += ingresos_mes
        total_deducciones += deducciones_mes

        resumen_meses.append(
            {
                "mes": mes,
                "ingresos": round(ingresos_mes, 2),
                "deducciones": round(deducciones_mes, 2),
                "cfdis": len(cfdis_mes),
            }
        )

    # ISR trimestral (acumulado)
    isr_trimestral = calcular_isr_estimado(total_ingresos, total_deducciones)

    return jsonify(
        {
            "trimestre": trimestre,
            "nombre": f"{trimestre} (Abr-Jun)" if q == 2 else trimestre,
            "meses": meses,
            "resumen_meses": resumen_meses,
            "totales": {
                "ingresos": round(total_ingresos, 2),
                "deducciones": round(total_deducciones, 2),
                "base_gravable": isr_trimestral["base_gravable"],
                "isr_estimado": isr_trimestral["isr_estimado"],
                "reserva_mensual": isr_trimestral["reserva_mensual"],
            },
            "tendencia": "↑" if total_ingresos > 0 else "→",
        }
    )


@app.route("/calculo-isr")
def calculo_isr():
    """Cálculo de ISR estimado con proyección"""
    cache = load_cache()

    # Obtener año
    anio = request.args.get("anio", str(date.today().year))

    # Calcular ingresos acumulados del año
    cfdis_anio = [
        c
        for c in cache.get("recibidos", [])
        if c.get("fecha_emision", "").startswith(anio) and c.get("efecto") == "I"
    ]

    ingresos_acumulados = sum(float(c.get("monto", 0)) for c in cfdis_anio)
    deducciones_acumuladas = sum(
        float(c.get("monto", 0)) for c in cfdis_anio if clasificar_gasto(c)["deducible"]
    )

    # Proyección anual (si vamos a mitad de año, proyectar)
    mes_actual = date.today().month if anio == str(date.today().year) else 12
    if mes_actual > 0:
        factor_proyeccion = 12 / mes_actual
        ingresos_proyectados = ingresos_acumulados * factor_proyeccion
        deducciones_proyectadas = deducciones_acumuladas * factor_proyeccion
    else:
        ingresos_proyectados = ingresos_acumulados
        deducciones_proyectadas = deducciones_acumuladas

    isr_actual = calcular_isr_estimado(ingresos_acumulados, deducciones_acumuladas)
    isr_proyectado = calcular_isr_estimado(
        ingresos_proyectados, deducciones_proyectadas
    )

    return jsonify(
        {
            "anio": anio,
            "mes_actual": mes_actual,
            "acumulado": {
                "ingresos": round(ingresos_acumulados, 2),
                "deducciones": round(deducciones_acumuladas, 2),
                "isr_estimado": isr_actual["isr_estimado"],
                "reserva_mensual": isr_actual["reserva_mensual"],
            },
            "proyeccion_anual": {
                "ingresos_estimados": round(ingresos_proyectados, 2),
                "deducciones_estimadas": round(deducciones_proyectadas, 2),
                "isr_estimado": isr_proyectado["isr_estimado"],
                "reserva_mensual": isr_proyectado["reserva_mensual"],
                "reserva_acumulada": round(
                    isr_proyectado["reserva_mensual"] * mes_actual, 2
                ),
            },
            "recomendacion": f"Reserva ${isr_proyectado['reserva_mensual']:,.2f} mensual para ISR",
        }
    )


# ─── Optimización Fiscal (Diferenciadores vs Konta) ───────────────────


@app.route("/optimize/suggestions")
def optimize_suggestions():
    """Sugerencias de optimización fiscal basadas en CFDIs reales"""
    from optimization_engine import OptimizationEngine
    from pathlib import Path

    user_id = request.args.get("user_id", "marco_test")
    data_dir = Path(f"/app/data/users/{user_id}")

    try:
        engine = OptimizationEngine(str(data_dir))
        result = engine.generate_report()
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error en optimización: {e}", exc_info=True)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/optimize/projection")
def optimize_projection():
    """Proyección ISR y detección de salto de tarifa"""
    from optimization_engine import OptimizationEngine
    from pathlib import Path

    user_id = request.args.get("user_id", "marco_test")
    data_dir = Path(f"/app/data/users/{user_id}")

    try:
        engine = OptimizationEngine(str(data_dir))
        result = engine.project_isr()

        # Agregar alerta si hay salto de tarifa
        proy = result.get("proyeccion_anual", {})
        diferencia = proy.get("diferencia", 0)

        alertas = []
        if diferencia > 0:
            alertas.append(
                {
                    "tipo": "salto_tarifa",
                    "titulo": "⚠️ Posible salto de tarifa",
                    "mensaje": f"Tu ISR real podría ser ${diferencia:,.2f} mayor que tu estimación actual. Considera aumentar deducciones antes de que cierre el año.",
                    "accion": "Revisa /optimize/suggestions para ver cómo reducir tu base gravable.",
                }
            )

        result["alertas"] = alertas
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error en proyección: {e}", exc_info=True)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/complementos-pendientes")
def complementos_pendientes():
    """Lista facturas PPD con saldo pendiente y vencimientos"""
    cache = load_cache()
    hoy = date.today()

    # Facturas PPD emitidas
    ppd_emitidas = [
        c
        for c in cache.get("emitidos", [])
        if c.get("metodo_pago") == "PPD" and c.get("estatus") == "1"
    ]

    pendientes = []
    for f in ppd_emitidas:
        total = float(f.get("monto", 0))
        pagado = float(f.get("total_pagado", 0))
        saldo = total - pagado

        if saldo > 0:
            # Calcular fecha límite para complemento (10 del mes siguiente)
            fecha_emision = f.get("fecha_emision", hoy.isoformat())
            try:
                fecha_pago = date.fromisoformat(fecha_emision)
                # Asumir que el pago se hizo el mismo mes, complemento vence 10 del siguiente
                if fecha_pago.month == 12:
                    fecha_limite = date(fecha_pago.year + 1, 1, 10)
                else:
                    fecha_limite = date(fecha_pago.year, fecha_pago.month + 1, 10)

                dias_restantes = (fecha_limite - hoy).days

                pendientes.append(
                    {
                        "uuid": f.get("uuid", ""),
                        "nombre_receptor": f.get("nombre_receptor", ""),
                        "rfc_receptor": f.get("rfc_receptor", ""),
                        "fecha_emision": fecha_emision,
                        "monto_total": total,
                        "pagado": pagado,
                        "saldo_pendiente": round(saldo, 2),
                        "fecha_limite_complemento": fecha_limite.isoformat(),
                        "dias_restantes": dias_restantes,
                        "urgencia": "critical"
                        if dias_restantes < 0
                        else ("high" if dias_restantes <= 3 else "normal"),
                    }
                )
            except:
                pass

    # Ordenar por urgencia
    pendientes.sort(key=lambda x: x["dias_restantes"])

    return jsonify(
        {
            "fecha_actual": hoy.isoformat(),
            "total_pendientes": len(pendientes),
            "total_saldo": round(sum(p["saldo_pendiente"] for p in pendientes), 2),
            "urgentes": len([p for p in pendientes if p["urgencia"] == "critical"]),
            "pendientes": pendientes,
        }
    )


@app.route("/clasificacion-gastos")
def clasificacion_gastos():
    """Análisis de gastos por categoría"""
    cache = load_cache()

    # Filtros de fecha
    mes = request.args.get("mes")  # YYYY-MM
    anio = request.args.get("anio", str(date.today().year))

    cfdis = cache.get("recibidos", [])

    # Filtrar por período
    if mes:
        cfdis = [c for c in cfdis if get_mes_fiscal(c.get("fecha_emision", "")) == mes]
    elif anio:
        cfdis = [c for c in cfdis if c.get("fecha_emision", "").startswith(anio)]

    # Clasificar todos los gastos
    categorias = {}
    for c in cfdis:
        clasif = clasificar_gasto(c)
        categoria = clasif["clasificacion"]

        if categoria not in categorias:
            categorias[categoria] = {
                "total": 0,
                "count": 0,
                "deducible": clasif["deducible"],
                "cfdis": [],
            }

        monto = float(c.get("monto", 0))
        categorias[categoria]["total"] += monto
        categorias[categoria]["count"] += 1
        categorias[categoria]["cfdis"].append(
            {
                "uuid": c.get("uuid", "")[:8] + "...",
                "monto": monto,
                "fecha": c.get("fecha_emision", ""),
            }
        )

    # Ordenar por monto
    categorias_ordenadas = dict(
        sorted(categorias.items(), key=lambda x: x[1]["total"], reverse=True)
    )

    # Totales
    total_deducible = sum(
        cat["total"] for cat in categorias.values() if cat["deducible"]
    )
    total_no_deducible = sum(
        cat["total"] for cat in categorias.values() if not cat["deducible"]
    )

    return jsonify(
        {
            "periodo": mes or anio,
            "total_categorias": len(categorias),
            "total_deducible": round(total_deducible, 2),
            "total_no_deducible": round(total_no_deducible, 2),
            "categorias": categorias_ordenadas,
        }
    )


@app.route("/emitir", methods=["POST"])
def emitir():
    """Register an invoice issued from SAT portal"""
    data = request.json or {}

    # Create invoice record
    invoice = {
        "uuid": data.get("uuid") or str(uuid.uuid4()).upper(),
        "nombre_emisor": data.get("nombre_emisor", "MARCO ARTURO MUÑOZ DEL TORO"),
        "rfc_emisor": data.get("rfc_emisor", "MUTM8610091NA"),
        "nombre_receptor": data.get("nombre_receptor", ""),
        "rfc_receptor": data.get("rfc_receptor", ""),
        "monto": float(data.get("monto", 0)),
        "fecha_emision": data.get("fecha_emision", date.today().isoformat()),
        "efecto": "I",
        "estatus": "1",
        "metodo_pago": data.get("metodo_pago", "PUE"),
        "forma_pago": data.get("forma_pago", "03"),
        "uso_cfdi": data.get("uso_cfdi", "G03"),
        "complementos": [],
        "total_pagado": 0
        if data.get("metodo_pago") == "PPD"
        else float(data.get("monto", 0)),
        "saldo_pendiente": float(data.get("monto", 0))
        if data.get("metodo_pago") == "PPD"
        else 0,
        "fecha_registro": datetime.now().isoformat(),
    }

    cache = load_cache()
    cache.setdefault("emitidos", []).append(invoice)
    save_cache(cache)

    return jsonify(
        {
            "status": "success",
            "message": "Factura registrada",
            "uuid": invoice["uuid"],
            "invoice": invoice,
        }
    )


@app.route("/timbrar", methods=["POST"])
def timbrar():
    """
    Genera y timbra un CFDI 4.0 usando el portal gratuito del SAT.
    """
    data = request.json or {}

    # Validar datos requeridos
    emisor = data.get("emisor", {})
    receptor = data.get("receptor", {})
    conceptos = data.get("conceptos", [])

    if not emisor.get("rfc") or not receptor.get("rfc") or not conceptos:
        return jsonify(
            {
                "status": "error",
                "message": "Faltan datos: emisor.rfc, receptor.rfc, conceptos",
            }
        ), 400

    try:
        from timbrado import TimbradoSAT, FacturaBuilder
        from pathlib import Path

        # Obtener vault dir del usuario
        rfc_emisor = emisor.get("rfc", "MUTM8610091NA")
        vault_dir = Path(DATA_DIR) / "users" / rfc_emisor.lower() / "vault"

        # Crear timbrador
        timbrador = TimbradoSAT(rfc=rfc_emisor, vault_dir=str(vault_dir))

        # Preparar datos de factura
        builder = FacturaBuilder()
        factura = builder.build_factura_base(
            emisor_rfc=emisor.get("rfc"),
            emisor_nombre=emisor.get("nombre", ""),
            receptor_rfc=receptor.get("rfc"),
            receptor_nombre=receptor.get("nombre", ""),
            conceptos=conceptos,
            total=data.get("total", 0),
            uso_cfdi=receptor.get("uso_cfdi", "G03"),
        )
        factura.update(
            {
                "forma_pago": data.get("forma_pago", "03"),
                "metodo_pago": data.get("metodo_pago", "PUE"),
                "serie": data.get("serie", "A"),
                "folio": data.get("folio", ""),
            }
        )

        # Timbrar
        resultado = timbrador.timbrar(factura)

        if resultado.get("status") == "success":
            # Guardar en cache de emitidos
            cache = load_cache()
            cache.setdefault("emitidos", []).append(
                {
                    "uuid": resultado.get("uuid"),
                    "rfc_emisor": rfc_emisor,
                    "rfc_receptor": receptor.get("rfc"),
                    "nombre_receptor": receptor.get("nombre", ""),
                    "total": data.get("total", 0),
                    "fecha_emision": datetime.now().isoformat(),
                    "xml": resultado.get("xml", ""),
                    "estatus": "1",
                }
            )
            save_cache(cache)

        return jsonify(resultado)

    except Exception as e:
        logger.error(f"Timbrado error: {e}", exc_info=True)
        return jsonify(
            {"status": "error", "message": f"Error en timbrado: {str(e)}"}
        ), 500


@app.route("/cancelar", methods=["POST"])
def cancelar():
    """
    Cancela un CFDI usando el servicio gratuito del SAT.
    """
    data = request.json or {}
    uuid = data.get("uuid", "")
    motivo = data.get("motivo", "02")

    if not uuid:
        return jsonify({"status": "error", "message": "UUID requerido"}), 400

    try:
        from timbrado import TimbradoSAT

        rfc_emisor = data.get("rfc_emisor", "MUTM8610091NA")
        vault_dir = Path(DATA_DIR) / "users" / rfc_emisor.lower() / "vault"

        timbrador = TimbradoSAT(rfc=rfc_emisor, vault_dir=str(vault_dir))
        resultado = timbrador.cancelar(uuid=uuid, motivo=motivo)

        if resultado.get("status") == "success":
            # Actualizar cache
            cache = load_cache()
            for cfdi_list in [cache.get("recibidos", []), cache.get("emitidos", [])]:
                for cfdi in cfdi_list:
                    if cfdi.get("uuid") == uuid:
                        cfdi["estatus"] = "0"
                        cfdi["fecha_cancelacion"] = datetime.now().isoformat()
            save_cache(cache)

        return jsonify(resultado)

    except Exception as e:
        logger.error(f"Cancelación error: {e}", exc_info=True)
        return jsonify(
            {"status": "error", "message": f"Error en cancelación: {str(e)}"}
        ), 500


@app.route("/estado-cfdi/<uuid>")
def estado_cfdi(uuid):
    """Consulta el estado de un CFDI en el SAT"""
    try:
        from timbrado import TimbradoSAT

        rfc = request.args.get("rfc", "MUTM8610091NA")
        vault_dir = Path(DATA_DIR) / "users" / rfc.lower() / "vault"

        timbrador = TimbradoSAT(rfc=rfc, vault_dir=str(vault_dir))
        resultado = timbrador.validar_estado_cfdi(uuid=uuid)

        return jsonify(resultado)

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/opinion")
def opinion_cumplimiento():
    connector = sat_connector()
    if not connector:
        return jsonify(
            {"status": "error", "message": "SAT connector no disponible"}
        ), 500
    try:
        if not connector._is_connected:
            connector.authenticate()
        opinion = connector.get_compliance_opinion()
        return jsonify(opinion)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/obligaciones")
def obligaciones():
    today = date.today()
    year, month, day = today.year, today.month, today.day

    obligations = []

    if day <= 17:
        days_to_17 = 17 - day
        urgency = (
            "critical" if days_to_17 <= 3 else ("high" if days_to_17 <= 7 else "normal")
        )
        obligations.append(
            {
                "id": "iva-isr-mensual",
                "titulo": f"IVA + ISR - {MONTH_NAMES[str(month).zfill(2)]} {year}",
                "tipo": "Mensual",
                "vence": f"{year}-{str(month).zfill(2)}-17",
                "dias_restantes": days_to_17,
                "urgencia": urgency,
                "descripcion": "Declaración mensual de IVA e ISR.",
                "accion": "Presenta antes del 17",
            }
        )
    else:
        next_m = month + 1 if month < 12 else 1
        next_y = year if month < 12 else year + 1
        obligations.append(
            {
                "id": "iva-isr-mensual",
                "titulo": f"IVA + ISR - {MONTH_NAMES[str(next_m).zfill(2)]} {next_y}",
                "tipo": "Mensual",
                "vence": f"{next_y}-{str(next_m).zfill(2)}-17",
                "dias_restantes": 47 - day,
                "urgencia": "normal",
                "descripcion": "Próxima declaración.",
            }
        )

    if month <= 4:
        days_to_april_30 = (date(year, 4, 30) - today).days
        if days_to_april_30 > 0:
            urgency = (
                "critical"
                if days_to_april_30 <= 7
                else ("high" if days_to_april_30 <= 21 else "normal")
            )
            obligations.append(
                {
                    "id": "declaracion-anual",
                    "titulo": f"Declaración Anual {year - 1}",
                    "tipo": "Anual",
                    "vence": f"{year}-04-30",
                    "dias_restantes": days_to_april_30,
                    "urgencia": urgency,
                }
            )

    urgency_order = {"critical": 0, "overdue": 1, "high": 2, "normal": 3, "low": 4}
    obligations.sort(key=lambda x: urgency_order.get(x["urgencia"], 5))

    return jsonify(
        {
            "rfc": "MUTM8610091NA",
            "fecha_actual": today.isoformat(),
            "obligaciones_pendientes": obligations,
            "resumen": {
                "criticas": sum(1 for o in obligations if o["urgencia"] == "critical"),
                "altas": sum(1 for o in obligations if o["urgencia"] == "high"),
                "total": len(obligations),
            },
        }
    )


@app.route("/complemento-pago", methods=["POST"])
def complemento_pago():
    """Generar complemento de pago para una factura PPD"""
    data = request.json or {}
    factura_uuid = data.get("factura_uuid")
    monto_pagado = data.get("monto_pagado", 0)
    fecha_pago = data.get("fecha_pago")
    forma_pago = data.get("forma_pago", "03")

    if not factura_uuid:
        return jsonify({"status": "error", "message": "UUID de factura requerido"}), 400

    cache = load_cache()
    emitidos = cache.get("emitidos", [])

    # Find the original invoice
    factura = next((c for c in emitidos if c.get("uuid") == factura_uuid), None)
    if not factura:
        return jsonify({"status": "error", "message": "Factura no encontrada"}), 404

    # Initialize complementos list if not exists
    if "complementos" not in factura:
        factura["complementos"] = []

    # Calculate total paid so far
    total_pagado = sum(
        c.get("monto_pagado", 0) for c in factura.get("complementos", [])
    )
    total_pagado += float(monto_pagado)

    # Create complemento record
    complemento = {
        "uuid": str(uuid.uuid4()).upper(),
        "tipo": "P",
        "factura_relacionada": factura_uuid,
        "monto_pagado": float(monto_pagado),
        "fecha_pago": fecha_pago,
        "forma_pago": forma_pago,
        "fecha_generacion": datetime.now().isoformat(),
        "estatus": "1",
    }

    factura["complementos"].append(complemento)

    # Update payment status
    factura["total_pagado"] = total_pagado
    factura["saldo_pendiente"] = float(factura.get("monto", 0)) - total_pagado

    save_cache(cache)

    return jsonify(
        {
            "status": "success",
            "message": "Complemento de pago generado",
            "complemento": complemento,
            "factura": {
                "uuid": factura_uuid,
                "total": factura.get("monto"),
                "pagado": total_pagado,
                "pendiente": factura["saldo_pendiente"],
            },
        }
    )


# Fix: remove duplicate __main__ blocks and keep only one at the end
@app.route("/declaracion/borrador/<tipo>")
def declaracion_borrador(tipo):
    """
    Genera borrador de declaracion en PDF o XML con datos pre-completados.
    tipo: 'ceros' | 'normal'
    Query:
      periodo=YYYY-MM (default: mes anterior)
      rfc=RFC_EMISOR
      regimen=612|626
      formato=pdf|xml (default: pdf)
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    periodo = request.args.get(
        "periodo", (date.today().replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    )
    rfc = request.args.get("rfc", "MUTM8610091NA")
    regimen = request.args.get("regimen", "612")
    formato = request.args.get("formato", "pdf").lower()

    cache = load_cache()
    recibidos = cache.get("recibidos", [])
    emitidos = cache.get("emitidos", [])
    cfdis = recibidos + emitidos

    total_ingresos = sum(
        float(c.get("monto", 0)) for c in cfdis if c.get("efecto") in ("I", "ingreso")
    )
    total_egresos = sum(
        float(c.get("monto", 0)) for c in cfdis if c.get("efecto") in ("E", "egreso")
    )
    isr_info = calcular_isr_estimado(total_ingresos, total_egresos, regimen)

    if formato == "xml" and tipo == "ceros":
        # Genera XML tipo presentacion pre-llenado (formato simplificado del SAT Diot)
        # Esto no es un XML oficial del SAT, pero organiza los datos segun formato entendible
        content = f"""<?xml version="1.0" encoding="UTF-8"?>
<BorradorDeclaracion version="1.0" tipo="{tipo}">
    <Contribuyente rfc="{rfc}" regimen="{regimen}"/>
    <Periodo>{periodo}</Periodo>
    <Ingresos total="{isr_info["ingresos"]:,.2f}"/>
    <Deducciones total="{isr_info["deducciones"]:,.2f}"/>
    <BaseGravable>{isr_info["base_gravable"]:,.2f}</BaseGravable>
    <ISRCalculado>{isr_info["isr_estimado"]:,.2f}</ISRCalculado>
    <Nota>ESTE ES UN BORRADOR GENERADO POR FISCOMIND. DEBES PRESENTAR EN EL PORTAL OFICIAL DEL SAT.</Nota>
</BorradorDeclaracion>"""
        return (
            content,
            200,
            {
                "Content-Type": "application/xml",
                "Content-Disposition": f"attachment; filename=borrador_declaracion_{tipo}_{periodo}.xml",
            },
        )

    # DEFAULT: PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=18,
    )
    elements = []
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=colors.HexColor("#533483"),
        spaceAfter=20,
    )
    elements.append(Paragraph("FiscoMind — Borrador de Declaración", title))
    elements.append(
        Paragraph(f"Contribuyente: {rfc}  |  Periodo: {periodo}", styles["Normal"])
    )
    elements.append(Spacer(1, 0.3 * inch))

    if tipo == "ceros":
        elements.append(
            Paragraph("<b>TIPO: DECLARACIÓN EN CEROS</b>", styles["Heading2"])
        )
        elements.append(
            Paragraph(
                "No se detectaron ingresos deducibles para este periodo.",
                styles["Normal"],
            )
        )
    else:
        elements.append(
            Paragraph(f"<b>TIPO: DECLARACIÓN NORMAL</b>", styles["Heading2"])
        )
    elements.append(Spacer(1, 0.2 * inch))

    summary_data = [
        ["Concepto", "Monto"],
        ["Total Ingresos", f"${isr_info['ingresos']:,.2f}"],
        ["Total Deducciones", f"${isr_info['deducciones']:,.2f}"],
        ["Base Gravable", f"${isr_info['base_gravable']:,.2f}"],
        ["ISR Calculado", f"${isr_info['isr_estimado']:,.2f}"],
        ["Reserva Mensual", f"${isr_info['reserva_mensual']:,.2f}"],
    ]
    table = Table(summary_data, colWidths=[3.5 * inch, 2.5 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#533483")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 14),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(table)

    elements.append(Spacer(1, 0.3 * inch))
    elements.append(
        Paragraph(
            "<b>INSTRUCCIONES PARA PRESENTAR:</b><br/>"
            + "1. Portal SAT &rarr; Declaraciones y Pagos<br/>"
            + "2. Selecciona el periodo correspondiente<br/>"
            + '3. Si es "ceros": selecciona declaración en ceros<br/>'
            + '4. Si es "normal": captura los montos de este borrador<br/>'
            + "5. Presenta y guarda el acuse",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(
        Paragraph(
            "<i>Este es un borrador informativo generado por FiscoMind. "
            "Verifica los datos antes de presentar ante el SAT.</i>",
            styles["Italic"],
        )
    )
    doc.build(elements)
    return (
        buffer.getvalue(),
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f"attachment; filename=borrador_declaracion_{tipo}_{periodo}.pdf",
        },
    )


@app.route("/simular", methods=["POST"])
def simular_escenario():
    """
    Simulador de escenarios fiscales "what-if"

    Escenarios soportados:
    - compra_activo: Compra de equipo (laptop, muebles) con depreciación
    - incremento_ingresos: Proyectar más facturación
    - incremento_deducciones: Gastos adicionales deducibles
    - honorarios: Calcular retenciones

    Request body:
    {
        "tipo": "compra_activo",
        "monto": 25000,
        "descripcion": "Laptop Dell XPS",
        "mes": "2026-05" (opcional, default: mes actual)
    }
    """
    cache = load_cache()
    data = request.json or {}

    tipo = data.get("tipo", "compra_activo")
    monto = float(data.get("monto", 0))
    descripcion = data.get("descripcion", "")
    mes_simulacion = data.get("mes", date.today().strftime("%Y-%m"))

    # Calcular situación actual (mes seleccionado)
    cfdis_mes = [
        c
        for c in cache.get("recibidos", [])
        if get_mes_fiscal(c.get("fecha_emision", "")) == mes_simulacion
        and c.get("efecto") == "I"
    ]

    ingresos_actuales = sum(float(c.get("monto", 0)) for c in cfdis_mes)
    deducciones_actuales = sum(
        float(c.get("monto", 0)) for c in cfdis_mes if clasificar_gasto(c)["deducible"]
    )

    isr_actual = calcular_isr_estimado(ingresos_actuales, deducciones_actuales)

    # Calcular escenario simulado
    if tipo == "compra_activo":
        # Depreciación inmediata para equipos de cómputo <= $25,000
        if monto <= 25000 and "computo" in descripcion.lower():
            deduccion_inmediata = monto  # Deducción inmediata
            depreciacion_anual = 0
            nota = "✅ Deducción inmediata (Art. 31 frac. XII LISR)"
        else:
            # Depreciación normal
            deduccion_inmediata = 0
            depreciacion_anual = monto * 0.30  # 30% anual típico
            nota = f"📊 Depreciación 30% anual: ${depreciacion_anual:,.2f}/año"

        nuevas_deducciones = deducciones_actuales + deduccion_inmediata

    elif tipo == "incremento_ingresos":
        nuevas_deducciones = deducciones_actuales
        ingresos_actuales += monto
        nota = f"📈 Ingresos adicionales: ${monto:,.2f}"

    elif tipo == "incremento_deducciones":
        nuevas_deducciones = deducciones_actuales + monto
        nota = f"✅ Gastos deducibles adicionales: ${monto:,.2f}"

    elif tipo == "honorarios":
        # Retenciones ISR e IVA para honorarios
        retencion_isr = monto * 0.10  # 10% ISR
        retencion_iva = monto * 0.1067  # 10.67% IVA (2/3 de 16%)
        nota = f"💰 Retenciones: ISR ${retencion_isr:,.2f}, IVA ${retencion_iva:,.2f}"
        nuevas_deducciones = deducciones_actuales

    else:
        return jsonify({"error": f"Tipo de escenario no soportado: {tipo}"}), 400

    # Calcular ISR nuevo
    isr_nuevo = calcular_isr_estimado(ingresos_actuales, nuevas_deducciones)

    # Diferencias
    ahorro_isr = isr_actual["isr_estimado"] - isr_nuevo["isr_estimado"]

    return jsonify(
        {
            "status": "success",
            "escenario": {
                "tipo": tipo,
                "descripcion": descripcion,
                "monto": monto,
                "mes": mes_simulacion,
            },
            "situacion_actual": {
                "ingresos": round(ingresos_actuales, 2),
                "deducciones": round(deducciones_actuales, 2),
                "base_gravable": isr_actual["base_gravable"],
                "isr_estimado": isr_actual["isr_estimado"],
                "reserva_mensual": isr_actual["reserva_mensual"],
            },
            "situacion_simulada": {
                "ingresos": round(ingresos_actuales, 2),
                "deducciones": round(nuevas_deducciones, 2),
                "base_gravable": isr_nuevo["base_gravable"],
                "isr_estimado": isr_nuevo["isr_estimado"],
                "reserva_mensual": isr_nuevo["reserva_mensual"],
            },
            "resultado": {
                "ahorro_isr": round(ahorro_isr, 2),
                "porcentaje_ahorro": round(
                    (ahorro_isr / isr_actual["isr_estimado"]) * 100, 1
                )
                if isr_actual["isr_estimado"] > 0
                else 0,
                "nota": nota,
            },
            "recomendacion": f"💡 {nota}\n\nAhorro estimado: ${ahorro_isr:,.2f} en ISR",
        }
    )


@app.route("/comparar", methods=["GET"])
def comparar_anios():
    """
    Comparativa año vs año

    Query params:
    - anio: Año a comparar (default: año anterior)
    """
    cache = load_cache()
    anio_comparar = request.args.get("anio", str(date.today().year - 1))
    anio_actual = str(date.today().year)

    mes_actual = date.today().month

    def get_resumen_anio(anio, mes_limite=None):
        """Obtener resumen hasta cierto mes"""
        cfdis = [
            c
            for c in cache.get("recibidos", [])
            if c.get("fecha_emision", "").startswith(anio) and c.get("efecto") == "I"
        ]

        if mes_limite:
            cfdis = [
                c
                for c in cfdis
                if int(c.get("fecha_emision", "2000-01-01").split("-")[1]) <= mes_limite
            ]

        ingresos = sum(float(c.get("monto", 0)) for c in cfdis)
        deducciones = sum(
            float(c.get("monto", 0)) for c in cfdis if clasificar_gasto(c)["deducible"]
        )

        isr = calcular_isr_estimado(ingresos, deducciones)

        return {
            "anio": anio,
            "meses": mes_limite or 12,
            "ingresos": round(ingresos, 2),
            "deducciones": round(deducciones, 2),
            "porcentaje_deducciones": round((deducciones / ingresos) * 100, 1)
            if ingresos > 0
            else 0,
            "isr_estimado": isr["isr_estimado"],
            "cfdis_count": len(cfdis),
        }

    # Resumen año actual hasta mes actual
    resumen_actual = get_resumen_anio(anio_actual, mes_actual)

    # Resumen año comparar hasta mismo mes
    resumen_comparar = get_resumen_anio(anio_comparar, mes_actual)

    # Calcular diferencias
    def calc_diff(actual, anterior, campo):
        if anterior == 0:
            return 0
        return round(((actual - anterior) / anterior) * 100, 1)

    diferencias = {
        "ingresos": calc_diff(
            resumen_actual["ingresos"], resumen_comparar["ingresos"], "ingresos"
        ),
        "deducciones": calc_diff(
            resumen_actual["deducciones"],
            resumen_comparar["deducciones"],
            "deducciones",
        ),
        "isr": calc_diff(
            resumen_actual["isr_estimado"], resumen_comparar["isr_estimado"], "isr"
        ),
        "porcentaje_deducciones": round(
            resumen_actual["porcentaje_deducciones"]
            - resumen_comparar["porcentaje_deducciones"],
            1,
        ),
    }

    # Alertas
    alertas = []
    if diferencias["porcentaje_deducciones"] < -2:
        alertas.append(
            "⚠️ Tu % de deducciones bajó. Podrías pagar más ISR del necesario."
        )
    if diferencias["ingresos"] > 20:
        alertas.append(
            "📈 Tus ingresos subieron significativamente. Considera estrategias de optimización."
        )
    if resumen_actual["porcentaje_deducciones"] < 10:
        alertas.append(
            "💡 Tus deducciones son bajas (<10%). Hay oportunidad de ahorrar ISR."
        )

    return jsonify(
        {
            "comparacion": {
                "anio_actual": anio_actual,
                "anio_comparar": anio_comparar,
                "meses_comparados": mes_actual,
            },
            "resumen_actual": resumen_actual,
            "resumen_comparar": resumen_comparar,
            "diferencias": diferencias,
            "alertas": alertas,
            "tendencia": "📈" if diferencias["ingresos"] > 0 else "📉",
        }
    )


# Exportación de datos REALES (archivos descargables)


@app.route("/export/<format>")
def export_data(format):
    """DESCARGA REAL DE ARCHIVOS: csv, pdf, xlsx"""
    user_id = request.args.get("user_id", "marco_test")

    cache = load_cache()
    cfdis = cache.get("recibidos", []) + cache.get("emitidos", [])

    # Filtro opcional: mes, año, tipo
    filtros = {
        k: request.args.get(k) for k in ["mes", "anio", "tipo"] if request.args.get(k)
    }
    if filtros.get("tipo"):
        cfdis = [c for c in cfdis if c.get("efecto") == filtros["tipo"]]
    if filtros.get("mes"):
        cfdis = [
            c for c in cfdis if get_mes_fiscal(c.get("fecha_emision")) == filtros["mes"]
        ]
    if filtros.get("anio"):
        cfdis = [
            c for c in cfdis if c.get("fecha_emision", "").startswith(filtros["anio"])
        ]

    if format == "csv":
        from export_tools import export_cfdis_to_csv

        csv_data = export_cfdis_to_csv(cfdis)
        return (
            csv_data,
            200,
            {
                "Content-Type": "text/csv",
                "Content-Disposition": f"attachment; filename=fiscomind_{user_id}_{datetime.now().strftime('%Y%m%d')}.csv",
            },
        )

    elif format == "pdf":
        from export_tools import generate_pdf_report

        pdf_bytes = generate_pdf_report(user_id=user_id, cfdis=cfdis)
        if pdf_bytes and isinstance(pdf_bytes, bytes) and len(pdf_bytes) > 100:
            return (
                pdf_bytes,
                200,
                {
                    "Content-Type": "application/pdf",
                    "Content-Disposition": f"attachment; filename=fiscomind_{user_id}_{datetime.now().strftime('%Y%m%d')}.pdf",
                },
            )
        else:
            return jsonify({"status": "error", "message": "PDF generation failed"}), 500

    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "CFDIs"
            headers = [
                "UUID",
                "Tipo",
                "Estatus",
                "Emisor",
                "RFC_Emisor",
                "Receptor",
                "RFC_Receptor",
                "Monto",
                "Fecha",
                "Uso CFDI",
                "Clasificación",
                "Deducible",
            ]
            ws.append(headers)
            for h in ws[1]:
                h.font = Font(bold=True)
                h.fill = PatternFill(
                    start_color="533483", end_color="533483", fill_type="solid"
                )
            for c in cfdis:
                clasif = clasificar_gasto(c)
                ws.append(
                    [
                        c.get("uuid", ""),
                        c.get("efecto", ""),
                        "Vigente" if c.get("estatus") == "1" else "Cancelado",
                        c.get("nombre_emisor", ""),
                        c.get("rfc_emisor", ""),
                        c.get("nombre_receptor", ""),
                        c.get("rfc_receptor", ""),
                        float(c.get("monto", 0)),
                        c.get("fecha_emision", ""),
                        c.get("uso_cfdi", ""),
                        clasif.get("clasificacion", ""),
                        "Sí" if clasif.get("deducible") else "No",
                    ]
                )
            ws.append([])
            ws.append(["RESUMEN"])
            ws.append(["Total Registros", len(cfdis)])
            ws.append(
                [
                    "Total Ingresos",
                    sum(
                        float(c.get("monto", 0))
                        for c in cfdis
                        if c.get("efecto") == "I"
                    ),
                ]
            )
            ws.append(
                [
                    "Total Egresos",
                    sum(
                        float(c.get("monto", 0))
                        for c in cfdis
                        if c.get("efecto") == "E"
                    ),
                ]
            )
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return (
                buffer.getvalue(),
                200,
                {
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": f"attachment; filename=fiscomind_{user_id}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                },
            )
        except ImportError:
            return jsonify(
                {"status": "error", "message": "openpyxl no instalado. Usa CSV o PDF."}
            ), 500

    else:
        return jsonify(
            {
                "status": "error",
                "message": f"Formato '{format}' no soportado. Use csv, pdf, xlsx",
            }
        ), 400


# Import Facturama routes
try:
    from facturama_routes import facturama_bp

    app.register_blueprint(facturama_bp, url_prefix="/facturama")
    logger.info("✅ Facturama routes registered")
except ImportError as e:
    logger.warning(f"⚠️ Facturama routes not loaded: {e}")


# ─── Regularización Fiscal ─────────────────────────────────────────


@app.route("/regularizacion", methods=["GET"])
def regularizacion_plan():
    """
    Genera plan de regularización fiscal completo.
    Analiza CFDIs, detecta períodos con/sin actividad, recomienda acciones.
    """
    from regularizacion_engine import (
        generar_plan_regularizacion,
        generar_pdf_regularizacion,
    )

    rfc = request.args.get("rfc", "MUTM8610091NA")
    formato = request.args.get("formato", "json")

    try:
        if formato == "pdf":
            pdf_bytes = generar_pdf_regularizacion(rfc=rfc, data_dir=str(DATA_DIR))
            if pdf_bytes and isinstance(pdf_bytes, bytes) and len(pdf_bytes) > 100:
                return (
                    pdf_bytes,
                    200,
                    {
                        "Content-Type": "application/pdf",
                        "Content-Disposition": f"attachment; filename=plan_regularizacion_{rfc}.pdf",
                    },
                )
            else:
                return jsonify(
                    {"status": "error", "message": "PDF generation failed"}
                ), 500

        plan = generar_plan_regularizacion(rfc=rfc, data_dir=str(DATA_DIR))
        return jsonify({"status": "success", "plan": plan})

    except Exception as e:
        logger.error(f"Regularizacion error: {e}", exc_info=True)
        return jsonify({"status": "error", "message": str(e)}), 500


# ─── ENDPOINTS ESPECÍFICOS CASO MARCO 2026 ─────────────────────────


@app.route("/marco/emparejar", methods=["GET"])
def marco_emparejar():
    """
    Emparejamiento bancario-fiscal completo del caso Marco 2026.
    Retorna: emitidos vigentes, duplicados, factura original, complemento de pago,
             movimientos bancarios clasificados, cálculo fiscal.
    """
    cache = load_cache()
    emitidos = cache.get("emitidos", [])
    recibidos = cache.get("recibidos", [])

    # Clasificar emitidos
    factura_original = None
    complemento = None
    duplicados = []
    otros_ingresos = []

    for c in emitidos:
        uuid = c.get("uuid", "")
        fecha = c.get("fecha_emision", "")
        monto = c.get("monto", 0)
        receptor = c.get("nombre_receptor", "")
        estatus = c.get("estatus", "")

        fact = {
            "uuid": uuid,
            "fecha": fecha,
            "monto": monto,
            "receptor": receptor,
            "rfc_receptor": c.get("rfc_receptor", ""),
            "estatus": "VIGENTE" if estatus == "1" else "CANCELADO",
        }

        if uuid == "DA4E4E23-C45D-4EC2-A057-A9606B65FA42":
            factura_original = fact
        elif uuid == "98A3AFB9-2E3E-46F3-82DF-D975A2CB5752":
            complemento = fact
        elif uuid in [
            "CF0741C5-D18E-40D2-BF4B-EF48A9A2E3D8",
            "26072D49-38FD-478A-89D9-7499206C29A4",
        ]:
            duplicados.append(fact)
        elif monto > 0:
            otros_ingresos.append(fact)

    # Calcular fiscal
    base_original = 57600.0 if factura_original else 0  # subtotal sin IVA
    iva_original = base_original * 0.16
    isr_original = base_original * 0.30
    total_impuesto = iva_original + isr_original

    # Movimientos bancarios del caso
    movimientos = [
        {
            "fecha": "2026-05-06",
            "tipo": "ingreso_tercero",
            "concepto": "SPEI Azteca / OSCAR ROBERTO TRUEBA FERNANDEZ",
            "monto": 950.00,
            "declarable": True,
        },
        {
            "fecha": "2026-04-22",
            "tipo": "transferencia_propia",
            "concepto": "Transferencia desde Fondeadora",
            "monto": 9000.00,
            "declarable": False,
        },
        {
            "fecha": "2026-04-20",
            "tipo": "transferencia_propia",
            "concepto": "Transferencia desde Fondeadora",
            "monto": 1000.00,
            "declarable": False,
        },
        {
            "fecha": "2026-04-20",
            "tipo": "transferencia_propia",
            "concepto": "Transferencia desde Fondeadora",
            "monto": 2869.00,
            "declarable": False,
        },
        {
            "fecha": "2026-04-06",
            "tipo": "transferencia_propia",
            "concepto": "SPEI NU MEXICO / Marco Arturo Muñoz Del Toro",
            "monto": 7040.00,
            "declarable": False,
        },
        {
            "fecha": "2026-04-30",
            "tipo": "gasto_deducible",
            "concepto": "GASOL GPO OCTANO",
            "monto": 1000.00,
            "declarable": False,
            "tiene_cfdi": None,
        },
        {
            "fecha": "2026-04-08",
            "tipo": "gasto_deducible",
            "concepto": "GASOL GPO OCTANO FLAMI",
            "monto": 1000.00,
            "declarable": False,
            "tiene_cfdi": None,
        },
        {
            "fecha": "2026-04-09",
            "tipo": "gasto_deducible",
            "concepto": "OPENAI 5.00 USD TC 17.50",
            "monto": 87.51,
            "declarable": False,
            "tiene_cfdi": None,
        },
    ]

    # Deducciones necesarias para llegar a cero
    deducciones_necesarias = base_original

    # Sugerencias de gastos faltantes
    sugerencias_gastos = [
        {
            "concepto": "Gasolina / transporte",
            "monto_estimado_mensual": 3000,
            "acumulado_3meses": 9000,
            "cfdi_requerido": True,
        },
        {
            "concepto": "Software / tecnología",
            "monto_estimado_mensual": 1500,
            "acumulado_3meses": 4500,
            "cfdi_requerido": True,
        },
        {
            "concepto": "Comidas / hospedaje negocio",
            "monto_estimado_mensual": 2000,
            "acumulado_3meses": 6000,
            "cfdi_requerido": True,
            "limite_ley": "No más del 10% de ingresos",
        },
        {
            "concepto": "Renta oficina / cowork",
            "monto_estimado_mensual": 5000,
            "acumulado_3meses": 15000,
            "cfdi_requerido": True,
        },
        {
            "concepto": "Servicios profesionales",
            "monto_estimado_mensual": 3000,
            "acumulado_3meses": 9000,
            "cfdi_requerido": True,
        },
    ]

    total_sugerido = sum(s["acumulado_3meses"] for s in sugerencias_gastos)

    return jsonify(
        {
            "status": "success",
            "caso": "marco_2026",
            "rfc": "MUTM8610091NA",
            "resumen": {
                "factura_original_ppd": factura_original,
                "complemento_pago": complemento,
                "duplicados_pendientes": duplicados,
                "otros_ingresos": otros_ingresos,
                "total_emitidos_vigentes": len(emitidos),
                "total_monto_vigente": sum(
                    c.get("monto", 0) for c in emitidos if c.get("estatus") == "1"
                ),
            },
            "fiscal_marzo_2026": {
                "base_gravable": base_original,
                "iva_trasladado": round(iva_original, 2),
                "isr_estimado": round(isr_original, 2),
                "total_impuesto": round(total_impuesto, 2),
                "deducciones_necesarias_para_cero": deducciones_necesarias,
            },
            "movimientos_bancarios": movimientos,
            "estrategia_deducciones": {
                "gastos_actuales_detectados": 2087.51,
                "gastos_faltantes_para_cero": deducciones_necesarias - 2087.51,
                "sugerencias_gastos": sugerencias_gastos,
                "total_sugerido_acumulado": total_sugerido,
                "factible_llegar_a_cero": total_sugerido >= deducciones_necesarias,
            },
            "advertencia": "Las deducciones deben ser reales, necesarias para la actividad, y documentadas con CFDIs. El SAT detecta patrones simulados.",
        }
    )


@app.route("/marco/deducir", methods=["GET"])
def marco_deducir():
    """
    Calculadora: cuánto deducir para minimizar ISR e IVA.
    Query: ?gastos=20000 (monto de gastos con CFDI que tienes)
    """
    gastos = float(request.args.get("gastos", 0))
    base_original = 57600.0
    iva_original = base_original * 0.16

    # ISR con deducciones
    base_isr = max(0, base_original - gastos)
    isr = base_isr * 0.30

    # IVA con acreditamiento
    iva_acreditable = gastos * 0.16
    iva_a_cargo = max(0, iva_original - iva_acreditable)

    # Total
    total = isr + iva_a_cargo

    return jsonify(
        {
            "status": "success",
            "escenario": f"Con ${gastos:,.2f} en gastos deducibles",
            "detalle": {
                "base_gravable_isr": round(base_isr, 2),
                "isr": round(isr, 2),
                "iva_trasladado_original": round(iva_original, 2),
                "iva_acreditable": round(iva_acreditable, 2),
                "iva_a_cargo": round(iva_a_cargo, 2),
                "total_impuestos": round(total, 2),
            },
            "ahorro_vs_sin_deducciones": round(26496 - total, 2),
            "mensaje": "Sin deducciones pagas ~$26,496. Cada $1,000 en gastos deducibles reduce ~$460 de impuestos.",
        }
    )


@app.route("/marco/marzo", methods=["GET"])
def marco_marzo():
    """Detalle fiscal del mes marzo 2026 para Marco."""
    return jsonify(
        {
            "status": "success",
            "mes": "2026-03",
            "rfc": "MUTM8610091NA",
            "factura_principal": {
                "uuid": "DA4E4E23-C45D-4EC2-A057-A9606B65FA42",
                "fecha": "2026-03-26",
                "receptor": "BENITTOS FOOD PARTY",
                "rfc_receptor": "BFP250829MN7",
                "subtotal": 57600.00,
                "iva": 9216.00,
                "total": 66616.00,
                "metodo_pago": "PPD",
                "tipo_comprobante": "I (Ingreso)",
            },
            "otros_emitidos_mes": [
                {
                    "uuid": "CF0741C5-D18E-40D2-BF4B-EF48A9A2E3D8",
                    "monto": 41669.85,
                    "nota": "DUPLICADO — cancelar",
                },
                {
                    "uuid": "26072D49-38FD-478A-89D9-7499206C29A4",
                    "monto": 45889.60,
                    "nota": "DUPLICADO — cancelar",
                },
            ],
            "obligaciones_mes": {
                "iva": {"base": 57600.00, "tasa": 0.16, "a_cargo": 9216.00},
                "isr": {"base": 57600.00, "tasa": 0.30, "estimado": 17280.00},
                "total_estimado": 26496.00,
            },
            "acciones": [
                "1. Verificar cancelación duplicados en SAT portal",
                "2. Juntar CFDIs de gastos deducibles del periodo",
                "3. Presentar declaración antes del 17 de mayo (si es mes actual)",
                "4. Guardar acuse",
            ],
        }
    )


@app.route("/marco/facilidades", methods=["GET"])
def marco_facilidades():
    """
    Facilidades de pago SAT para Marco.
    Calcula pagos diferidos, recargos, y opciones realistas.
    """
    base_original = 57600.0
    iva_original = base_original * 0.16
    isr_original = base_original * 0.30
    total_deuda = iva_original + isr_original  # $26,496

    opciones = []

    # Opción 1: Pago único inmediato
    opciones.append(
        {
            "opcion": 1,
            "nombre": "Pago único inmediato",
            "plazo_meses": 0,
            "pago_mensual": total_deuda,
            "recargos": 0,
            "total": total_deuda,
            "requiere": "Liquidez inmediata de $26,496",
            "ventaja": "Sin recargos, cierra inmediatamente",
        }
    )

    # Opción 2: Pago a 6 meses
    recargos_6 = total_deuda * 0.015 * 6  # 1.5% mensual aprox
    total_6 = total_deuda + recargos_6
    opciones.append(
        {
            "opcion": 2,
            "nombre": "Facilidades de pago a 6 meses",
            "plazo_meses": 6,
            "pago_mensual": round(total_6 / 6, 2),
            "recargos": round(recargos_6, 2),
            "total": round(total_6, 2),
            "requiere": "Solicitud en SAT portal + compromiso de pago",
            "ventaja": "Pagos mensuales de ~$4,600, menos presión inmediata",
        }
    )

    # Opción 3: Pago a 12 meses
    recargos_12 = total_deuda * 0.015 * 12
    total_12 = total_deuda + recargos_12
    opciones.append(
        {
            "opcion": 3,
            "nombre": "Facilidades de pago a 12 meses",
            "plazo_meses": 12,
            "pago_mensual": round(total_12 / 12, 2),
            "recargos": round(recargos_12, 2),
            "total": round(total_12, 2),
            "requiere": "Solicitud en SAT portal + garantía o aval",
            "ventaja": "Pagos mensuales de ~$2,500, máxima flexibilidad",
        }
    )

    # Opción 4: Condonación parcial (si aplica programa)
    opciones.append(
        {
            "opcion": 4,
            "nombre": "Programa de condonación/regularización",
            "plazo_meses": "Variable",
            "pago_mensual": "Variable",
            "recargos": 0,
            "total": "Desconocido - depende del programa vigente",
            "requiere": "Verificar si hay programa activo en SAT.gob.mx",
            "ventaja": "Posible condonación de recargos y multas",
            "nota": "Requiere asesoría profesional. No siempre disponible.",
        }
    )

    return jsonify(
        {
            "status": "success",
            "deuda_base": round(total_deuda, 2),
            "detalle_deuda": {
                "iva": round(iva_original, 2),
                "isr": round(isr_original, 2),
            },
            "opciones": opciones,
            "como_solicitar": {
                "portal_sat": "sat.gob.mx → Servicios → Facilidades de pago",
                "telefono": "55 627 22 728",
                "requerimientos": [
                    "Opinión de cumplimiento actualizada",
                    "No tener adeudos en otro proceso",
                    "Presentar declaración del periodo primero",
                ],
            },
            "advertencia": "Las facilidades de pago generan recargos mensuales (~1.47% actualización + recargo). Evalúa si te conviene pagar de contado vs. a plazos.",
        }
    )


@app.route("/declarar/proceso", methods=["GET"])
def declarar_proceso():
    """
    Cronograma estrategico para declarar marzo-abril-mayo 2026.
    Considera: fechas vencidas, deducciones, facilidades de pago, y riesgo SAT.
    """
    from datetime import date, timedelta

    hoy = date.today()

    # Fechas clave IVA e ISR para PFAE (regimen general)
    # IVA: mes siguiente dias 1-17
    # ISR provisional: mes siguiente dias 1-17
    # ISR anual: abril del siguiente año

    fechas_vencimiento = {
        "marzo_2026": {
            "periodo": "2026-03",
            "iva_vence": "2026-04-17",
            "isr_vence": "2026-04-17",
            "dias_retraso_iva": max(0, (hoy - date(2026, 4, 17)).days),
            "dias_retraso_isr": max(0, (hoy - date(2026, 4, 17)).days),
            "estado": "VENCIDO" if hoy > date(2026, 4, 17) else "AL DIA",
        },
        "abril_2026": {
            "periodo": "2026-04",
            "iva_vence": "2026-05-17",
            "isr_vence": "2026-05-17",
            "dias_retraso_iva": max(0, (hoy - date(2026, 5, 17)).days),
            "dias_retraso_isr": max(0, (hoy - date(2026, 5, 17)).days),
            "estado": "VENCIDO" if hoy > date(2026, 5, 17) else "AL DIA",
        },
        "mayo_2026": {
            "periodo": "2026-05",
            "iva_vence": "2026-06-17",
            "isr_vence": "2026-06-17",
            "dias_retraso_iva": 0,
            "dias_retraso_isr": 0,
            "estado": "AL DIA",
        },
    }

    # Multas por extemporaneidad
    multa_base = 1083  # aprox UMAs
    recargo_mensual = 0.0147  # 1.47% mensual

    # Calcular recargos para marzo
    meses_retraso_marzo = max(1, (hoy - date(2026, 4, 17)).days // 30)
    recargo_marzo_iva = 9216 * recargo_mensual * meses_retraso_marzo
    recargo_marzo_isr = 17280 * recargo_mensual * meses_retraso_marzo
    multa_extemporanea = multa_base * 2  # una por IVA, una por ISR

    # Calcular recargos para abril (si aplica)
    meses_retraso_abril = max(0, (hoy - date(2026, 5, 17)).days // 30)
    recargo_abril = 0
    if meses_retraso_abril > 0:
        iva_abril = 950 * 0.16  # ~152
        isr_abril = 950 * 0.30  # ~285
        recargo_abril = (iva_abril + isr_abril) * recargo_mensual * meses_retraso_abril

    # Escenarios
    escenarios = [
        {
            "nombre": "ESCENARIO A: Presentar todo HOY con deducciones minimas",
            "descripcion": "Presentar marzo, abril, mayo hoy. Sin deducciones adicionales.",
            "acciones": [
                "1. Presentar IVA marzo: $9,216 + recargos",
                "2. Presentar ISR marzo: $17,280 + recargos",
                "3. Presentar abril y mayo",
                "4. Pagar o solicitar facilidades",
            ],
            "costo_marzo": {
                "iva_base": 9216,
                "isr_base": 17280,
                "recargos_iva": round(recargo_marzo_iva, 2),
                "recargos_isr": round(recargo_marzo_isr, 2),
                "multas": round(multa_extemporanea, 2),
                "total_marzo": round(
                    9216
                    + 17280
                    + recargo_marzo_iva
                    + recargo_marzo_isr
                    + multa_extemporanea,
                    2,
                ),
            },
            "costo_abril": round(152 + 285 + recargo_abril, 2),
            "costo_mayo": 0,
            "total": round(
                9216
                + 17280
                + recargo_marzo_iva
                + recargo_marzo_isr
                + multa_extemporanea
                + 152
                + 285
                + recargo_abril,
                2,
            ),
            "ventaja": "Cierra todo inmediatamente",
            "riesgo": "Alto pago inmediato si no tienes liquidez",
        },
        {
            "nombre": "ESCENARIO B: Juntar deducciones 15 dias, presentar 31 mayo",
            "descripcion": "Usar tiempo restante de mayo para recopilar CFDIs de gastos. Presentar todo junto.",
            "acciones": [
                "1. Recopilar TODOS CFDIs de gastos hasta 31 mayo",
                "2. Presentar marzo con deducciones maximas",
                "3. Presentar abril y mayo",
                "4. Calcular si compensacion IVA a favor aplica",
            ],
            "costo_marzo": {
                "nota": "Variable segun gastos recopilados",
                "sin_deducciones": 26496,
                "con_5000_deducciones": 24196,
                "con_10000_deducciones": 21896,
                "con_25000_deducciones": 14996,
            },
            "recargos_abril_mayo": "Igual que escenario A si abril ya vencio",
            "ventaja": "Minimiza impuestos si logras juntar deducciones",
            "riesgo": "Si no juntas suficientes, pagas igual + recargos extra por esperar",
        },
        {
            "nombre": "ESCENARIO C: Facilidades de pago + presentar puntual",
            "descripcion": "Presentar declaraciones al dia pero pagar a plazos via SAT",
            "acciones": [
                "1. Presentar todas las declaraciones correctamente",
                "2. Solicitar facilidades de pago SAT",
                "3. Pagar a 6 o 12 meses",
            ],
            "costo": {
                "deuda_base": 26496,
                "recargos_facilidades_6meses": round(26496 * 0.015 * 6, 2),
                "total_6meses": round(26496 * 1.09, 2),
                "pago_mensual_6meses": round(26496 * 1.09 / 6, 2),
            },
            "ventaja": "Presentas al dia (menos multas), pagas a plazos",
            "riesgo": "Recargos mensuales adicionales",
        },
    ]

    # Recomendacion final
    recomendacion = ""
    if hoy < date(2026, 5, 17):
        recomendacion = "ABRIL AUN NO VENCE. Prioridad: presentar abril hoy, juntar deducciones para marzo, presentar marzo antes de junio."
    elif hoy < date(2026, 6, 17):
        recomendacion = "ABRIL VENCIDO, MAYO AUN AL DIA. Prioridad: presentar mayo inmediatamente (en ceros si no hay actividad), juntar deducciones para marzo-abril, presentar todo antes de 17 junio."
    else:
        recomendacion = "TODO VENCIDO. Prioridad: regularizar completo, solicitar facilidades de pago, juntar deducciones para minimizar."

    return jsonify(
        {
            "status": "success",
            "fecha_hoy": hoy.isoformat(),
            "resumen_vencimientos": fechas_vencimiento,
            "multas_estimadas": {
                "multa_base_extemporaneidad": multa_base,
                "recargo_mensual": f"{recargo_mensual * 100}%",
                "meses_retraso_marzo": meses_retraso_marzo,
                "recargo_estimado_marzo": round(
                    recargo_marzo_iva + recargo_marzo_isr, 2
                ),
            },
            "escenarios": escenarios,
            "recomendacion_urgente": recomendacion,
            "acciones_inmediatas": [
                "1. Verificar si abril 2026 ya se presento (o vencio)",
                "2. Si no se presento abril: URGENTE presentar hoy",
                "3. Recopilar CFDIs de gastos deducibles del periodo",
                "4. Calcular escenario mas favorable",
                "5. Presentar marzo con deducciones (o sin ellas si no alcanzas)",
                "6. Evaluar facilidades de pago si el monto es alto",
            ],
        }
    )


# ─── Upload y Análisis de Documentos ────────────────────────────────────


@app.route("/upload/estado-cuenta", methods=["POST"])
def upload_estado_cuenta():
    """
    Recibe texto de estado de cuenta (o CSV) y analiza movimientos fiscalmente.
    Input: JSON con {"texto": "copia-pega del estado", "banco": "Fondeadora", "mes": "2026-03"}
    Output: Clasificación completa de cada movimiento con impacto fiscal.
    """
    data = request.json or {}
    texto = data.get("texto", "")
    banco = data.get("banco", "")
    mes = data.get("mes", "")
    rfc = data.get("rfc", "MUTM8610091NA")

    if not texto or len(texto) < 50:
        return jsonify(
            {
                "status": "error",
                "message": "Texto requerido. Pega el estado de cuenta completo.",
            }
        ), 400

    try:
        import time as _time

        start = _time.time()
        from document_analyzer import EstadoCuentaAnalyzer

        analyzer = EstadoCuentaAnalyzer()
        resultado = analyzer.analizar(texto, banco=banco, mes=mes)

        # Agregar metadatos
        resultado.update(
            {
                "rfc": rfc,
                "timestamp": datetime.now().isoformat(),
                "tiempo_procesamiento": round(_time.time() - start, 3),
            }
        )

        return jsonify(resultado)

    except Exception as e:
        logger.error(f"Error analizando estado de cuenta: {e}", exc_info=True)
        return jsonify(
            {"status": "error", "message": f"Error al analizar: {str(e)}"}
        ), 500


@app.route("/analyze/situacion", methods=["POST"])
def analyze_situacion():
    """
    Motor de inteligencia fiscal completo.
    Recibe: CFDIs + movimientos + emitidos. Retorna: estrategia, riesgo, acciones.
    """
    data = request.json or {}

    try:
        from fiscal_intelligence import FiscalIntelligenceEngine

        engine = FiscalIntelligenceEngine(
            rfc=data.get("rfc", "MUTM8610091NA"),
            regimen=data.get("regimen", "PFAE-general"),
        )
        engine.context = data.get("contexto", {})

        resultado = engine.analyze(
            cfdis=data.get("cfdis", []),
            movimientos=data.get("movimientos", []),
            facturas_emitidas=data.get("emitidos", []),
            historial_declaraciones=data.get("historial", {}),
        )

        return jsonify(resultado)

    except Exception as e:
        logger.error(f"Error analizando situación fiscal: {e}", exc_info=True)
        return jsonify({"status": "error", "message": f"Error: {str(e)}"}), 500


@app.route("/marco/historial", methods=["GET"])
def marco_historial():
    """
    Historial fiscal completo 2022-2026.
    Detecta meses con ingresos que IMPIDEN declarar en ceros.
    Muestra lo que tu contadora NO ve en el SAT.
    """
    cache = load_cache()
    recibidos = cache.get("recibidos", [])
    emitidos = cache.get("emitidos", [])

    # Agrupar recibidos por mes
    meses_data = {}
    for c in recibidos:
        fecha = c.get("fecha_emision", "")
        if len(fecha) >= 7:
            mes = fecha[:7]
            efecto = c.get("efecto", "")
            monto = float(c.get("monto", 0))
            if mes not in meses_data:
                meses_data[mes] = {
                    "ingresos": 0,
                    "egresos": 0,
                    "pagos": 0,
                    "cfdis_count": 0,
                    "emitidos": 0,
                }
            meses_data[mes][
                {"I": "ingresos", "E": "egresos", "P": "pagos"}.get(efecto, "otros")
            ] += monto
            meses_data[mes]["cfdis_count"] += 1

    # Agregar emitidos por mes
    for c in emitidos:
        fecha = c.get("fecha_emision", "")
        if len(fecha) >= 7 and c.get("estatus") == "1":
            mes = fecha[:7]
            monto = float(c.get("monto", 0))
            if mes not in meses_data:
                meses_data[mes] = {
                    "ingresos": 0,
                    "egresos": 0,
                    "pagos": 0,
                    "cfdis_count": 0,
                    "emitidos": 0,
                }
            meses_data[mes]["emitidos"] += monto

    # Evaluar cada mes
    resultado_meses = []
    meses_criticos = []
    total_ingresos = 0
    total_emitidos = 0

    for mes in sorted(meses_data.keys()):
        d = meses_data[mes]
        total_mes = d["ingresos"] + d["egresos"] + d["pagos"]
        hay_emitidos = d["emitidos"] > 0

        puede_ceros = total_mes < 1000 and not hay_emitidos

        mes_info = {
            "mes": mes,
            "ingresos_recibidos": round(d["ingresos"], 2),
            "egresos_recibidos": round(d["egresos"], 2),
            "pagos_recibidos": round(d["pagos"], 2),
            "emitidos": round(d["emitidos"], 2),
            "total_actividad": round(total_mes + d["emitidos"], 2),
            "cfdis_count": d["cfdis_count"],
            "puede_declarar_ceros": puede_ceros,
            "riesgo_ceros": "BAJO" if puede_ceros else "ALTO",
            "explicacion": (
                "Puede ser ceros si son gastos personales"
                if puede_ceros
                else "Tiene actividad económica detectada. NO declarar ceros."
            ),
        }

        resultado_meses.append(mes_info)
        total_ingresos += d["ingresos"]
        total_emitidos += d["emitidos"]

        if not puede_ceros:
            meses_criticos.append(mes)

    # Evaluación global
    total_meses = len(resultado_meses)
    meses_ok = sum(1 for m in resultado_meses if m["puede_declarar_ceros"])
    meses_nok = total_meses - meses_ok

    return jsonify(
        {
            "status": "success",
            "rfc": "MUTM8610091NA",
            "analisis_periodo": "2022-01 a 2026-02",
            "resumen_global": {
                "total_meses_con_cfdis": total_meses,
                "meses_que_pueden_ser_ceros": meses_ok,
                "meses_con_actividad_real": meses_nok,
                "meses_criticos": meses_criticos,
                "total_ingresos_recibidos": round(total_ingresos, 2),
                "total_emitidos": round(total_emitidos, 2),
            },
            "advertencia_contadora": "Muchas contadoras NO revisan CFDIs recibidos en el SAT. Solo ven emitidos. Los CFDIs recibidos PROVEEN que hubo actividad económica.",
            "meses_detalle": resultado_meses,
            "recomendacion_final": (
                f"De {total_meses} meses con CFDIs, {meses_nok} tienen actividad real. "
                f"NO puedes declarar ceros en: {', '.join(meses_criticos[:10])}. "
                f"La estrategia de 'todos ceros' es RIESGOSA con la IA del SAT 2026."
            ),
        }
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
